import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from selenium import webdriver
from time import time
import shutil

input_folder = 'D:\\VHDFC_INTRA\\Input\\'
output_folder = 'D:\\VHDFC_INTRA\\Output\\'

try:
    list_file = os.listdir(input_folder)

    if len(list_file) == 1:
        excel_file_path = input_folder + list_file[0]

        excel_df = pd.read_excel(excel_file_path, engine='openpyxl')

        wb = load_workbook(excel_file_path)
        source = wb.active
        ws = wb.worksheets[0]

        row_count = len(excel_df.index)
        for index in range(row_count):

            row_of_df = excel_df.iloc[index]
            web_addr = row_of_df['Links']

            chrome_option = webdriver.ChromeOptions()
            chrome_option.add_experimental_option('useAutomationExtension', False)
            chrome_driver = r'D:\VHDFC_INTRA\Chrome_driver\chromedriver.exe'
            browser = webdriver.Chrome(chrome_driver)

            print(web_addr)
            start_time = time()
            browser.get(str(web_addr))
            end_time = time()
            time_to_load = end_time - start_time
            browser.quit()
            if time_to_load <= 5.0:

                ws.cell(row=index+2, column=1).fill = PatternFill(fgColor="00008000", fill_type="solid")
                ws.cell(row=index+2, column=2).value = str(time_to_load)
                ws.cell(row=index+2, column=4).value = str(True)
                wb.save(excel_file_path)

                print("Less than 5 sec",time_to_load)

            elif time_to_load > 5.0:

                ws.cell(row=index+2, column=1).fill = PatternFill(fgColor="00FF0000", fill_type="solid")
                ws.cell(row=index+2, column=2).value = str(time_to_load)
                ws.cell(row=index+2, column=3).value = str(True)
                wb.save(excel_file_path)

                print("Grater than equal to 5 sec",time_to_load)

        shutil.move(excel_file_path,output_folder+str(list_file[0]))


    else:
        print("Input folder have more than one")

except Exception as exp:
    print(Exception.args)
    print(exp)


